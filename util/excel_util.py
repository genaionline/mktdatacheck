import pandas as pd
import configparser
import os
import warnings
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import numpy as np

class ExcelUtil:
    @staticmethod
    def _fill_unmerged_values(df, column_name):
        """
        Fill empty values in a column with the next non-empty value that is surrounded by empty values.
        This handles cases where Excel cells should have been merged but weren't.
        
        Args:
            df (pd.DataFrame): Input DataFrame
            column_name (str): Name of the column to process
            
        Returns:
            pd.DataFrame: DataFrame with filled values
        """
        # Create a copy to avoid modifying the original
        df = df.copy()
        
        # Get the series we're working with
        series = df[column_name]
        
        # Find indices where values are not empty/null
        non_empty_idx = series[series.notna()].index
        
        if len(non_empty_idx) == 0:
            return df
        
        # Process each non-empty value
        for i in range(len(non_empty_idx)):
            current_idx = non_empty_idx[i]
            current_value = series[current_idx]
            
            # If this is not the last non-empty value
            if i < len(non_empty_idx) - 1:
                next_idx = non_empty_idx[i + 1]
                
                # Check if the next non-empty value is surrounded by empty values
                if next_idx > current_idx + 1:  # There are empty values between
                    # Fill all empty values between current and next with current_value
                    df.loc[current_idx:next_idx-1, column_name] = current_value
            
            # Handle the last segment
            if i == len(non_empty_idx) - 1:
                # Fill all remaining empty values with the last non-empty value
                df.loc[current_idx:, column_name] = current_value
        
        return df

    @staticmethod
    def extract_data(excel_file):
        """
        Extract data from Excel file based on configuration in sg.config.ini
        
        Args:
            excel_file (str): Path to the Excel file
            
        Returns:
            pd.DataFrame: Processed DataFrame with required columns
        """
        # Read configuration
        config = configparser.ConfigParser()
        config_path = os.path.join('config', 'sg.config.ini')
        config.read(config_path)
        
        # Get Excel settings
        worksheet_name = config.get('excel', 'worksheet_name')
        content_start_row = int(config.get('excel', 'content_start_row'))
        content_end_row = int(config.get('excel', 'content_end_row'))
        
        # Get column settings
        channel_col = config.get('allowed_metrics_names', 'channel_name_col')
        metrics_col = config.get('allowed_metrics_names', 'metrics_name_col')
        
        # Get allowed metrics combinations
        allowed_metrics = []
        for key in config['allowed_metrics_names']:
            if key.startswith('acq_') or key.startswith('rep_'):
                channel, metric = config.get('allowed_metrics_names', key).split(',')
                allowed_metrics.append((channel.strip(), metric.strip()))
        
        # Get column ranges
        lya_start = config.get('last_year_actual', 'start_col_name')
        lya_end = config.get('last_year_actual', 'end_col_name')
        cya_start = config.get('current_year_actual', 'start_col_name')
        cya_end = config.get('current_year_actual', 'end_col_name')
        cyt_start = config.get('current_year_target', 'start_col_name')
        cyt_end = config.get('current_year_target', 'end_col_name')
        
        # Load workbook for merge cell handling
        wb = load_workbook(excel_file)
        ws = wb[worksheet_name]
        
        # Handle merged cells in channel column
        merged_ranges = []
        for merged_cell in ws.merged_cells.ranges:
            if channel_col in str(merged_cell):
                start_row = merged_cell.min_row
                end_row = merged_cell.max_row
                value = ws[f"{channel_col}{start_row}"].value
                ws.unmerge_cells(str(merged_cell))
                for row in range(start_row, end_row + 1):
                    ws[f"{channel_col}{row}"] = value
        
        # Save temporary file if we unmerged cells
        temp_file = None
        if merged_ranges:
            temp_file = "temp_unmerged.xlsx"
            wb.save(temp_file)
            excel_file = temp_file
        
        try:
            # Read Excel file
            df = pd.read_excel(
                excel_file,
                sheet_name=worksheet_name,
                skiprows=content_start_row - 1,
                nrows=content_end_row - content_start_row + 1
            )
            
            # Validate metrics combinations
            channel_col_idx = column_index_from_string(channel_col) - 1  # Convert to 0-based index
            metrics_col_idx = column_index_from_string(metrics_col) - 1
            channel_col_name = df.columns[channel_col_idx]
            metrics_col_name = df.columns[metrics_col_idx]
            
            # Handle unmerged cells in the channel column
            df = ExcelUtil._fill_unmerged_values(df, channel_col_name)
            
            for idx, row in df.iterrows():
                channel = row[channel_col_name]
                metric = row[metrics_col_name]
                if (channel, metric) not in allowed_metrics:
                    warnings.warn(f"Warning: Invalid metrics combination found - Channel: {channel}, Metric: {metric}")
            
            # Select required columns using column indices
            lya_start_idx = column_index_from_string(lya_start) - 1
            lya_end_idx = column_index_from_string(lya_end) - 1
            cya_start_idx = column_index_from_string(cya_start) - 1
            cya_end_idx = column_index_from_string(cya_end) - 1
            cyt_start_idx = column_index_from_string(cyt_start) - 1
            cyt_end_idx = column_index_from_string(cyt_end) - 1
            
            cols_to_keep = [
                df.columns[channel_col_idx],  # Channel column
                df.columns[metrics_col_idx],  # Metrics column
                *df.columns[lya_start_idx:lya_end_idx + 1],  # LYA columns
                *df.columns[cya_start_idx:cya_end_idx + 1],  # CYA columns
                *df.columns[cyt_start_idx:cyt_end_idx + 1]   # CYT columns
            ]
            
            df = df[cols_to_keep]
            
            return df
            
        finally:
            # Clean up temporary file if created
            if temp_file and os.path.exists(temp_file):
                os.remove(temp_file)
